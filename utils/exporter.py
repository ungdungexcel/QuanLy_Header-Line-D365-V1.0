# -*- coding: utf-8 -*-
"""Xuat file Excel Head/Line cho D365."""

from __future__ import annotations

import io
import re
from datetime import datetime

import pandas as pd

from config import (
    D365_HEADER_COLUMNS,
    D365_LINE_COLUMNS,
    D365_LINE_REQUIRED_COLUMNS,
    EXPORT_DATETIME_FORMAT,
)
import config as _config

# Cot an danh dau dong tach VAT (bo truoc khi ghi Excel)
VAT_SPLIT_HL_COL = "__vat_split_hl"

_DATE_COLUMN_PATTERN = re.compile(
    r"date|datetime|ngay|ngày|ship|receipt|invoice|(?:^|[_\s])time|giờ|gio",
    re.IGNORECASE,
)


def _parse_datetime(value) -> datetime | None:
    """Chuyen gia tri ve datetime; ho tro chuoi ngay VN va Excel serial."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()

    text = str(value).strip()
    if not text or text.lower() in {"nan", "nat", "none"}:
        return None

    if re.match(r"^\d{4}-\d{2}-\d{2}", text):
        parsed = pd.to_datetime(text, errors="coerce")
    else:
        parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    if pd.notna(parsed):
        return parsed.to_pydatetime()

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def format_datetime_value(value, *, use_now_if_empty: bool = False) -> str:
    """Dinh dang ngay: dd/mm/yyyy."""
    dt = _parse_datetime(value)
    if dt is None:
        if use_now_if_empty:
            return datetime.now().strftime(EXPORT_DATETIME_FORMAT)
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ""
        text = str(value).strip()
        return "" if text.lower() in {"nan", "nat", "none"} else text
    return dt.strftime(EXPORT_DATETIME_FORMAT)


def _is_date_column(col_name: str) -> bool:
    return bool(_DATE_COLUMN_PATTERN.search(str(col_name)))


def format_dataframe_for_export(df: pd.DataFrame) -> pd.DataFrame:
    """Chuan hoa cot ngay truoc khi xuat Excel."""
    if df.empty:
        return df.copy()

    out = df.copy()
    for col in out.columns:
        if str(col).startswith("__"):
            continue
        if _is_date_column(col):
            out[col] = out[col].map(format_datetime_value)
    return out


def _format_customer_account(value) -> str:
    """Giu nguyen Store code, bao gom so 0 dau (vd: 00232)."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, (int,)):
        return str(value)
    if isinstance(value, float) and value == int(value):
        # Da mat so 0 dau neu doc bang so — chi con phan nguyen
        return str(int(value))
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    # "232.0" tu Excel float -> "232"
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    return text


# Cot Header can giu dang text (so 0 dau)
_TEXT_EXPORT_COLUMNS = {
    "Customer account",
    "Invoice account",
    "Customer requisition",
    "Sales order",
    "Số SO#",
}


def _format_vat_display(value) -> str:
    """Hien thi thue VAT dang phan tram (08l/10l -> 8%/10%)."""
    text = str(value).strip() if value is not None else ""
    if not text or text.lower() in {"nan", "none"}:
        return ""
    m = re.match(r"^0*(\d+)\s*%?\s*[lL]?$", text)
    if m:
        return f"{int(m.group(1))}%"
    return text


def _line_vat_values(df_lines: pd.DataFrame, df_header: pd.DataFrame) -> pd.Series:
    """Lay ma VAT theo dong Line; thieu thi map tu Header theo SO."""
    if "vat" in df_lines.columns:
        vat = _series_text(df_lines, "vat")
    else:
        vat = pd.Series([""] * len(df_lines), index=df_lines.index, dtype=str)

    if vat.ne("").all() or df_header is None or df_header.empty or "vat" not in df_header.columns:
        return vat.map(_format_vat_display)

    vat_by_so = {
        str(r["so_number"]).strip(): _format_vat_display(r["vat"])
        for _, r in df_header.iterrows()
    }
    if "temp_so_number" in df_lines.columns:
        keys = df_lines["temp_so_number"].astype(str).str.strip()
    else:
        keys = df_lines["so_number"].astype(str).str.strip() if "so_number" in df_lines.columns else pd.Series([""] * len(df_lines))
    from_header = keys.map(vat_by_so).fillna("")
    filled = vat.where(vat.ne(""), from_header)
    return filled.map(_format_vat_display)


def _format_vat_purchaser_name(row: pd.Series) -> str:
    """Ten CH + so PO."""
    store_name = str(row.get("store_name", "")).strip()
    po_number = str(row.get("po_number", "")).strip()
    if store_name and po_number:
        return f"{store_name} {po_number}"
    return store_name or po_number


def _format_warning_notes(row: pd.Series) -> str:
    """Ghi chu tach SO, canh bao VAT va Memo tu PO."""
    parts: list[str] = []
    notes = str(row.get("notes", "")).strip()
    memo = str(row.get("memo", "")).strip()
    if notes:
        parts.append(notes)
    if memo and memo not in notes:
        parts.append(memo)
    return " | ".join(parts)


def _pick_header_date(row: pd.Series, *fields: str) -> str:
    """Lay ngay tu cac truong uu tien, dinh dang dd/mm/yyyy."""
    for field in fields:
        if field not in row.index:
            continue
        value = row.get(field)
        if value is None or (isinstance(value, float) and pd.isna(value)):
            continue
        if str(value).strip():
            return format_datetime_value(value)
    return ""


def _series_text(df: pd.DataFrame, col: str) -> pd.Series:
    """Lay cot text; neu thieu cot thi tra series rong."""
    if col not in df.columns:
        return pd.Series([""] * len(df), index=df.index, dtype=str)
    out = df[col].map(
        lambda v: "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v).strip()
    )
    return out.replace({"nan": "", "None": "", "NaT": ""})


def _site_warehouse_for_header(df_header: pd.DataFrame, df_lines: pd.DataFrame) -> tuple[pd.Series, pd.Series]:
    """Lay Site/Warehouse tu PO (Header); neu thieu tung dong thi lay tu Line theo SO."""
    site = _series_text(df_header, "site")
    warehouse = _series_text(df_header, "warehouse")

    need_site = site.eq("").any() and "site" in df_lines.columns
    need_wh = warehouse.eq("").any() and "warehouse" in df_lines.columns
    if not (need_site or need_wh) or "so_number" not in df_lines.columns:
        return site, warehouse

    lines = df_lines.copy()
    # uu tien map theo SO tam neu co (sau khi ghép D365)
    so_key = "temp_so_number" if "temp_so_number" in lines.columns else "so_number"
    if need_site:
        lines["_site"] = _series_text(lines, "site")
    if need_wh:
        lines["_warehouse"] = _series_text(lines, "warehouse")

    def _first_nonempty(series: pd.Series) -> str:
        for value in series:
            text = str(value).strip()
            if text and text.lower() not in {"nan", "none"}:
                return text
        return ""

    grouped = lines.groupby(lines[so_key].astype(str).str.strip(), sort=False)
    header_so = df_header["so_number"].astype(str).str.strip()
    if need_site:
        site_map = grouped["_site"].agg(_first_nonempty)
        from_lines = header_so.map(site_map).fillna("").astype(str)
        site = site.where(site.ne(""), from_lines)
    if need_wh:
        wh_map = grouped["_warehouse"].agg(_first_nonempty)
        from_lines = header_so.map(wh_map).fillna("").astype(str)
        warehouse = warehouse.where(warehouse.ne(""), from_lines)
    return site, warehouse


def fill_missing_site_warehouse(
    target: pd.DataFrame,
    source: pd.DataFrame,
    *,
    key_candidates: tuple[str, ...] = (
        "Sales order",
        "Số SO#",
        "so_number",
        "temp_so_number",
        "Customer requisition",
        "po_number",
    ),
    overwrite: bool = False,
) -> pd.DataFrame:
    """Bo sung Site/Warehouse tu source (PO) vao target.

    overwrite=True: uu tien gia tri tu source (file PO), ghi de len target.
    overwrite=False: chi dien khi target dang trong.
    """
    out = target.copy()
    if out.empty or source.empty:
        return out

    src = source.copy()
    if "Site" not in src.columns and "site" in src.columns:
        src["Site"] = src["site"]
    if "Warehouse" not in src.columns and "warehouse" in src.columns:
        src["Warehouse"] = src["warehouse"]

    if "Site" not in out.columns:
        out["Site"] = ""
    if "Warehouse" not in out.columns:
        out["Warehouse"] = ""

    site_vals = _series_text(src, "Site" if "Site" in src.columns else "site")
    wh_vals = _series_text(src, "Warehouse" if "Warehouse" in src.columns else "warehouse")
    if site_vals.eq("").all() and wh_vals.eq("").all():
        return out

    # Gop map tu moi cot khoa co gia tri (tranh dung cot Số SO# rong)
    site_map: dict[str, str] = {}
    wh_map: dict[str, str] = {}
    for key in key_candidates:
        if key not in src.columns:
            continue
        keys = src[key].astype(str).str.strip().replace({"nan": "", "None": ""})
        for k, s, w in zip(keys, site_vals, wh_vals):
            if not k:
                continue
            if s and k not in site_map:
                site_map[k] = s
            if w and k not in wh_map:
                wh_map[k] = w

    if not site_map and not wh_map:
        return out

    cur_site = out["Site"].astype(str).str.strip().replace({"nan": "", "None": ""})
    cur_wh = out["Warehouse"].astype(str).str.strip().replace({"nan": "", "None": ""})

    mapped_site = pd.Series([""] * len(out), index=out.index, dtype=str)
    mapped_wh = pd.Series([""] * len(out), index=out.index, dtype=str)
    for key in key_candidates:
        if key not in out.columns:
            continue
        keys = out[key].astype(str).str.strip().replace({"nan": "", "None": ""})
        hit_site = keys.map(site_map).fillna("")
        hit_wh = keys.map(wh_map).fillna("")
        mapped_site = mapped_site.where(mapped_site.ne(""), hit_site)
        mapped_wh = mapped_wh.where(mapped_wh.ne(""), hit_wh)

    if overwrite:
        # Uu tien Site/Warehouse tu file PO
        out["Site"] = mapped_site.where(mapped_site.ne(""), cur_site)
        out["Warehouse"] = mapped_wh.where(mapped_wh.ne(""), cur_wh)
    else:
        out["Site"] = cur_site.where(cur_site.ne(""), mapped_site)
        out["Warehouse"] = cur_wh.where(cur_wh.ne(""), mapped_wh)
    return out


def _vat_split_so_set(df_header: pd.DataFrame) -> set[str]:
    """Tap SO tam can highlight (SO tach them do khac thue — khong gom SO chinh cung thue)."""
    if df_header is None or df_header.empty or "so_number" not in df_header.columns:
        return set()
    flag_col = "vat_split_hl" if "vat_split_hl" in df_header.columns else "vat_split"
    if flag_col not in df_header.columns:
        return set()
    mask = df_header[flag_col].fillna(False).astype(bool)
    return set(df_header.loc[mask, "so_number"].astype(str).str.strip())


def _vat_split_po_set(df_header: pd.DataFrame) -> set[str]:
    """PO cua cac SO can highlight (chi SO tach them)."""
    if df_header is None or df_header.empty or "po_number" not in df_header.columns:
        return set()
    flag_col = "vat_split_hl" if "vat_split_hl" in df_header.columns else None
    if flag_col is None:
        return set()
    mask = df_header[flag_col].fillna(False).astype(bool)
    return set(df_header.loc[mask, "po_number"].astype(str).str.strip())


def apply_vat_split_row_flags(
    df: pd.DataFrame,
    df_header: pd.DataFrame,
    mapping: dict[str, str] | None = None,
    *,
    for_line: bool = False,
) -> pd.DataFrame:
    """Danh dau dong Header/Line cua SO tach them do khac thue.

    SO chinh (cung thue / Sales order dau) — khong danh dau.
    """
    out = df.copy()
    if out.empty:
        out[VAT_SPLIT_HL_COL] = []
        return out

    split_temps = _vat_split_so_set(df_header)
    split_d365: set[str] = set()
    if mapping:
        split_d365 = {str(mapping[t]).strip() for t in split_temps if t in mapping and mapping[t]}

    if not split_temps and not split_d365:
        out[VAT_SPLIT_HL_COL] = False
        return out

    flags: list[bool] = []
    for _, row in out.iterrows():
        so_temp = str(row.get("Số SO#", row.get("so_number", ""))).strip()
        so_d365 = str(row.get("Sales order", "")).strip()
        temp_so_line = str(row.get("temp_so_number", "")).strip()

        if for_line:
            flagged = (
                (temp_so_line and temp_so_line in split_temps)
                or (so_d365 and so_d365 in split_d365)
                or (so_d365 and so_d365 in split_temps)
                or (so_temp and so_temp in split_temps)
            )
        else:
            flagged = (so_temp and so_temp in split_temps) or (so_d365 and so_d365 in split_d365)
        flags.append(bool(flagged))

    out[VAT_SPLIT_HL_COL] = flags
    return out


def _attach_vat_split_flags(
    d365_header: pd.DataFrame,
    d365_line: pd.DataFrame,
    df_header: pd.DataFrame,
    df_lines: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Highlight Header + Line cua SO tach them (khac thue); SO chinh cung thue thi khong."""
    header = d365_header.copy()
    line = d365_line.copy()

    flag_col = "vat_split_hl" if df_header is not None and "vat_split_hl" in df_header.columns else "vat_split"
    if df_header is None or df_header.empty or flag_col not in df_header.columns:
        header[VAT_SPLIT_HL_COL] = False
        line[VAT_SPLIT_HL_COL] = False
        return header, line

    split_map = {
        str(r["so_number"]).strip(): bool(r[flag_col])
        for _, r in df_header.iterrows()
    }
    header[VAT_SPLIT_HL_COL] = (
        header["Số SO#"].astype(str).str.strip().map(split_map).fillna(False).astype(bool)
    )

    if "temp_so_number" in df_lines.columns:
        keys = df_lines["temp_so_number"].astype(str).str.strip()
    elif "so_number" in df_lines.columns:
        keys = df_lines["so_number"].astype(str).str.strip()
    else:
        keys = pd.Series([""] * len(line))

    line_flags = keys.map(split_map).fillna(False).astype(bool)
    if len(line_flags) == len(line):
        line[VAT_SPLIT_HL_COL] = line_flags.to_numpy()
    else:
        line[VAT_SPLIT_HL_COL] = (
            line["Sales order"].astype(str).str.strip().map(split_map).fillna(False).astype(bool)
        )
    return header, line


def build_d365_files(df_header: pd.DataFrame, df_lines: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Chuyen df noi bo sang format cot Header/Line Template (theo mau xuat D365)."""
    now_text = datetime.now().strftime(EXPORT_DATETIME_FORMAT)
    site_vals, warehouse_vals = _site_warehouse_for_header(df_header, df_lines)

    d365_header = pd.DataFrame(
        {
            "Số SO#": _series_text(df_header, "so_number").to_numpy(),
            "Sales order": "",
            "Status": "",
            "Requested ship date": df_header.apply(
                lambda r: _pick_header_date(r, "order_date", "delivery_date"), axis=1
            ).to_numpy(),
            "Requested receipt date": df_header.apply(
                lambda r: _pick_header_date(r, "delivery_date", "order_date"), axis=1
            ).to_numpy(),
            "VAT Invoice date": now_text,
            "Customer account": df_header["customer"].map(_format_customer_account).to_numpy(),
            "Invoice account": "",
            "Delivery address name": "",
            "Pool": "",
            "VAT purchaser name": df_header.apply(_format_vat_purchaser_name, axis=1).to_numpy(),
            "Site": site_vals.to_numpy(),
            "Warehouse": warehouse_vals.to_numpy(),
            "Customer requisition": _series_text(df_header, "po_number").to_numpy(),
            "Customer reference": df_header.apply(_format_warning_notes, axis=1).to_numpy(),
            "Mode of delivery": "",
        }
    )
    d365_header = d365_header.reindex(columns=_config.D365_HEADER_COLUMNS).fillna("")

    line_site = _series_text(df_lines, "site")
    line_wh = _series_text(df_lines, "warehouse")
    if line_site.eq("").any() or line_wh.eq("").any():
        # Lay theo SO tam neu Line da doi sang SO D365
        if "temp_so_number" in df_lines.columns:
            temp_keys = df_lines["temp_so_number"].astype(str).str.strip()
            site_by_temp = dict(zip(d365_header["Số SO#"].astype(str), d365_header["Site"]))
            wh_by_temp = dict(zip(d365_header["Số SO#"].astype(str), d365_header["Warehouse"]))
            fill_site = temp_keys.map(site_by_temp).fillna("")
            fill_wh = temp_keys.map(wh_by_temp).fillna("")
        else:
            so_keys = df_lines["so_number"].astype(str).str.strip()
            site_by_so = dict(zip(d365_header["Số SO#"].astype(str), d365_header["Site"]))
            wh_by_so = dict(zip(d365_header["Số SO#"].astype(str), d365_header["Warehouse"]))
            fill_site = so_keys.map(site_by_so).fillna("")
            fill_wh = so_keys.map(wh_by_so).fillna("")
        line_site = line_site.where(line_site.ne(""), fill_site)
        line_wh = line_wh.where(line_wh.ne(""), fill_wh)

    unit_vals = (
        _series_text(df_lines, "uom").replace("", "Ea").to_numpy()
        if "uom" in df_lines.columns
        else ["Ea"] * len(df_lines)
    )
    text_vals = (
        _series_text(df_lines, "product_name").to_numpy()
        if "product_name" in df_lines.columns
        else _series_text(df_lines, "notes").to_numpy()
    )

    d365_line = pd.DataFrame(
        {
            "Company": "",
            "Line status": "",
            "Site": line_site.to_numpy(),
            "Warehouse": line_wh.to_numpy(),
            "Sales order": _series_text(df_lines, "so_number").to_numpy(),
            "External item number": "",
            "Item number": _series_text(df_lines, "item").to_numpy(),
            "Delivery address name": "",
            "Text": text_vals,
            "Unit": unit_vals,
            "Quantity": df_lines["quantity"].to_numpy() if "quantity" in df_lines.columns else "",
            "Unit price": (
                _series_text(df_lines, "unit_price").to_numpy()
                if "unit_price" in df_lines.columns
                else ""
            ),
            "Net amount": "",
            "Requested receipt date": df_lines.apply(
                lambda r: _pick_header_date(r, "delivery_date", "order_date"), axis=1
            ).to_numpy(),
            "Thuế VAT": _line_vat_values(df_lines, df_header).to_numpy(),
        }
    )
    d365_line = d365_line.reindex(columns=_config.D365_LINE_COLUMNS).fillna("")

    d365_header, d365_line = _attach_vat_split_flags(d365_header, d365_line, df_header, df_lines)
    return format_dataframe_for_export(d365_header), format_dataframe_for_export(d365_line)


def apply_sales_order_to_header(d365_header: pd.DataFrame, mapping: dict[str, str]) -> pd.DataFrame:
    """Dien cot Sales order tu map SO tam -> SO D365 (giu nguyen format Header mau)."""
    out = d365_header.copy()
    if "Số SO#" not in out.columns:
        return out
    hl = out[VAT_SPLIT_HL_COL] if VAT_SPLIT_HL_COL in out.columns else None
    out["Sales order"] = out["Số SO#"].astype(str).str.strip().map(mapping).fillna("")
    out = out.reindex(columns=_config.D365_HEADER_COLUMNS).fillna("")
    if hl is not None:
        out[VAT_SPLIT_HL_COL] = hl.to_numpy() if hasattr(hl, "to_numpy") else hl
    return out


def build_d365_export_filename(kind: str = "Header", suffix: str = "") -> str:
    """Ten file xuat: D365_Header_ddMMyyyy_HHmmss.xlsx"""
    stamp = datetime.now().strftime("%d%m%Y_%H%M%S")
    mid = f"_{suffix}" if suffix else ""
    return f"D365_{kind}{mid}_{stamp}.xlsx"


def _style_line_header(worksheet, columns: list[str]) -> None:
    """Dinh dang dong tieu de Line: bat buoc = vang/do, con lai = xanh/trang."""
    from openpyxl.styles import Alignment, Font, PatternFill

    fill_required = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_optional = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font_required = Font(color="FF0000", bold=True)
    font_optional = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        if col_name in D365_LINE_REQUIRED_COLUMNS:
            cell.fill = fill_required
            cell.font = font_required
        else:
            cell.fill = fill_optional
            cell.font = font_optional
        cell.alignment = align


def _style_vat_split_rows(worksheet, flags: list[bool], n_cols: int) -> None:
    """Dong tach VAT: nen vang + chu do dam."""
    from openpyxl.styles import Font, PatternFill

    fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    font = Font(color="FF0000", bold=True)
    for i, flagged in enumerate(flags):
        if not flagged:
            continue
        excel_row = i + 2  # bo qua header
        for col_idx in range(1, n_cols + 1):
            cell = worksheet.cell(row=excel_row, column=col_idx)
            cell.fill = fill
            cell.font = font


def _pop_vat_split_flags(df: pd.DataFrame) -> tuple[pd.DataFrame, list[bool] | None]:
    if VAT_SPLIT_HL_COL not in df.columns:
        return df, None
    flags = [
        bool(v) and str(v).strip().lower() not in {"", "0", "false", "nan", "none"}
        for v in df[VAT_SPLIT_HL_COL].tolist()
    ]
    return df.drop(columns=[VAT_SPLIT_HL_COL]), flags


def _apply_text_column_format(worksheet, columns: list[str], n_rows: int) -> None:
    """Ep dinh dang Text (@) de Excel khong mat so 0 dau (Customer account...)."""
    for col_idx, col_name in enumerate(columns, start=1):
        if str(col_name).strip() not in _TEXT_EXPORT_COLUMNS:
            continue
        for row_idx in range(2, n_rows + 2):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                cell.value = ""
            else:
                cell.value = str(cell.value)
            cell.number_format = "@"


def to_excel_bytes(*sheets: tuple[str, pd.DataFrame]) -> bytes:
    """Ghi nhieu sheet vao bytes Excel — cot ngay da dinh dang dd/mm/yyyy."""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, df in sheets:
            export_df = format_dataframe_for_export(df)
            export_df, hl_flags = _pop_vat_split_flags(export_df)
            # Bo cot an neu con sot
            drop_cols = [c for c in export_df.columns if str(c).startswith("__")]
            if drop_cols:
                export_df = export_df.drop(columns=drop_cols)
            # Ep cot Customer account... sang chuoi truoc khi ghi
            for col in export_df.columns:
                if str(col).strip() in _TEXT_EXPORT_COLUMNS:
                    export_df[col] = export_df[col].map(
                        lambda v: "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v)
                    )
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)

            worksheet = writer.sheets[sheet_name]
            if sheet_name == "Line":
                _style_line_header(worksheet, list(export_df.columns))
            if hl_flags:
                _style_vat_split_rows(worksheet, hl_flags, len(export_df.columns))
            _apply_text_column_format(worksheet, list(export_df.columns), len(export_df))
            for col_idx, col_name in enumerate(export_df.columns, start=1):
                if not _is_date_column(col_name):
                    continue
                for row_idx in range(2, len(export_df) + 2):
                    worksheet.cell(row=row_idx, column=col_idx).number_format = "DD/MM/YYYY"
    return buffer.getvalue()
